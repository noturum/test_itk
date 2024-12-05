from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import render
from openpyxl import load_workbook
from .models import TableData
from django.core import serializers


class UploadDataView(APIView):
    def post(self, request):
        file = request.FILES.get("file")

        if not file:
            return Response(
                {"error": "file not find"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                tech = row[3] or ""
                coord = row[2].split(",")
                TableData.objects.create(
                    ne=row[0],
                    address=row[1],
                    latitude=coord[0],
                    longitude=coord[1],
                    gsm=tech.find("gsm") > 0,
                    umts=tech.find("umts") > 0,
                    lte=tech.find("lte") > 0,
                    status=row[4],
                )
            return Response(
                {"message": "File loaded successfully"}, status=status.HTTP_201_CREATED
            )
        except Exception:
            return Response({"message": "Bad file"}, status=status.HTTP_400_BAD_REQUEST)


def get_json(request):
    obj = TableData.objects.all()
    data = serializers.serialize("json", obj)
    return HttpResponse(data)


def data_view(request):
    data_object = TableData.objects.all()
    return render(request, "table.html", {"data": data_object})
